"""Excel metadata importer and exporter."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from models import MetadataField, MetadataProject, MetadataRecord
from schema_manager import fields_from_headers
from utils import slugify, write_csv

try:
    import openpyxl
    from openpyxl import Workbook
except Exception:  # noqa: BLE001
    openpyxl = None
    Workbook = None  # type: ignore[assignment]


@dataclass
class ImportErrorItem:
    row: int
    column: str
    message: str


@dataclass
class ExcelPreview:
    headers: list[str]
    rows: list[list[str]]
    suggested_mapping: dict[str, str]
    unknown_headers: list[str]
    errors: list[ImportErrorItem] = field(default_factory=list)


class ExcelImporter:
    def _check_dependency(self) -> None:
        if openpyxl is None:
            raise RuntimeError("openpyxl is not installed. Please run: pip install openpyxl")

    def preview(self, path: Path, project: MetadataProject, max_rows: int = 100) -> ExcelPreview:
        self._check_dependency()
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)  # type: ignore[union-attr]
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(v).strip() if v is not None else "" for v in next(rows_iter)]
        except StopIteration:
            return ExcelPreview([], [], {}, [])
        mapping, unknown = self.suggest_mapping(headers, project)
        rows: list[list[str]] = []
        for i, row in enumerate(rows_iter, start=1):
            if i > max_rows:
                break
            rows.append(["" if v is None else str(v) for v in row[: len(headers)]])
        wb.close()
        return ExcelPreview(headers=headers, rows=rows, suggested_mapping=mapping, unknown_headers=unknown)

    def suggest_mapping(self, headers: list[str], project: MetadataProject) -> tuple[dict[str, str], list[str]]:
        mapping: dict[str, str] = {}
        unknown: list[str] = []
        schema = project.schema
        lookup: dict[str, str] = {}
        for f in schema.fields:
            for key in {f.field_id, f.xml_tag, f.label_zh, f.label_en}:
                if key:
                    lookup[str(key).strip().lower()] = f.field_id
        for h in headers:
            if not h:
                continue
            fid = lookup.get(h.strip().lower()) or lookup.get(slugify(h))
            if fid:
                mapping[h] = fid
            else:
                unknown.append(h)
        return mapping, unknown

    def import_file(
        self,
        path: Path,
        project: MetadataProject,
        mapping: dict[str, str] | None = None,
        add_unknown_to_schema: bool = False,
        max_rows: int | None = None,
    ) -> tuple[int, list[ImportErrorItem]]:
        self._check_dependency()
        preview = self.preview(path, project, max_rows=5)
        headers = preview.headers
        if not project.schema.fields:
            project.schema = fields_from_headers(headers)
        mapping = mapping or preview.suggested_mapping
        if add_unknown_to_schema:
            for header in headers:
                if header and header not in mapping:
                    fid = slugify(header)
                    project.schema.add_field(MetadataField(field_id=fid, label_zh=header, label_en=header, xml_tag=fid))
                    mapping[header] = fid
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)  # type: ignore[union-attr]
        ws = wb.active
        errors: list[ImportErrorItem] = []
        count = 0
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_index == 1:
                continue
            if max_rows is not None and count >= max_rows:
                break
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            record = MetadataRecord()
            for col_index, header in enumerate(headers):
                if col_index >= len(row):
                    continue
                field_id = mapping.get(header)
                if not field_id:
                    continue
                value = row[col_index]
                if value is not None:
                    record.set_value(field_id, str(value))
            # text_id 或 record_id 可作为记录ID，避免重复。
            candidate = record.get_value("record_id") or record.get_value("text_id")
            if candidate:
                record.record_id = project.ensure_unique_record_id(candidate)
            project.add_record(record)
            count += 1
        wb.close()
        return count, errors

    def export_records_csv(self, project: MetadataProject, path: Path) -> None:
        fieldnames = ["record_id", "record_type"] + project.schema.field_ids()
        rows = [r.to_flat_dict(project.schema) for r in project.records]
        write_csv(path, rows, fieldnames)

    def export_records_excel(self, project: MetadataProject, path: Path) -> None:
        self._check_dependency()
        wb = Workbook()  # type: ignore[operator]
        ws = wb.active
        ws.title = "metadata_records"
        headers = ["record_id", "record_type"] + project.schema.field_ids()
        ws.append(headers)
        for record in project.records:
            row = record.to_flat_dict(project.schema)
            ws.append([row.get(h, "") for h in headers])
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 10), 40)
        wb.save(path)
