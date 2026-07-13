# -*- coding: utf-8 -*-
"""Import previously exported WebLens links without using a database."""
from __future__ import annotations

import csv
from datetime import datetime
import re
from pathlib import Path
import xml.etree.ElementTree as ET
from typing import Iterable

from .collector import SearchRecord

URL_RE = re.compile(r"https?://[^\s<>\]）)\"']+", re.I)


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def make_record(row: dict, rank: int = 0) -> SearchRecord | None:
    link = str(row.get("link") or row.get("url") or row.get("URL") or row.get("final_url") or "").strip()
    if not link:
        return None
    rec = SearchRecord(
        collected_at=str(row.get("collected_at") or row.get("Collected") or now_iso()),
        query=str(row.get("query") or row.get("Query") or "imported"),
        search_vertical=str(row.get("search_vertical") or row.get("vertical") or "imported"),
        shard_start=str(row.get("shard_start") or ""),
        shard_end=str(row.get("shard_end") or ""),
        page=int_safe(row.get("page"), 0),
        rank=int_safe(row.get("rank"), rank),
        title=str(row.get("title") or row.get("Title") or link),
        link=link,
        source=str(row.get("source") or row.get("Source") or ""),
        published_time=str(row.get("published_time") or row.get("Published") or row.get("date") or ""),
        snippet=str(row.get("snippet") or row.get("Snippet") or ""),
        search_url=str(row.get("search_url") or ""),
        language_lr=str(row.get("language_lr") or ""),
        country_cr=str(row.get("country_cr") or ""),
        search_engine=str(row.get("search_engine") or row.get("engine") or ""),
        source_filter=str(row.get("source_filter") or row.get("baidu_source_filter") or ""),
        sort_mode=str(row.get("sort_mode") or row.get("baidu_sort") or ""),
        site_limit=str(row.get("site_limit") or ""),
        actual_domain=str(row.get("actual_domain") or ""),
        query_raw=str(row.get("query_raw") or ""),
        date_filter_type=str(row.get("date_filter_type") or ""),
        date_start=str(row.get("date_start") or ""),
        date_end=str(row.get("date_end") or ""),
        start_ts=str(row.get("start_ts") or ""),
        end_ts=str(row.get("end_ts") or ""),
        baidu_gpc=str(row.get("baidu_gpc") or ""),
    )
    for field in [
        "content_status", "content_word_count", "content_quality_score", "content_error",
        "content_extraction_method", "content_cleaning_scheme",
        "raw_html_path", "raw_text_path", "clean_text_path", "metadata_path", "metadata_excel_path",
    ]:
        if row.get(field) not in (None, ""):
            setattr(rec, field, row.get(field))
    return rec


def int_safe(value, default=0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except Exception:
        return default


def import_records(path: str | Path) -> list[SearchRecord]:
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix == ".csv":
        return import_csv(p)
    if suffix == ".xlsx":
        return import_xlsx(p)
    if suffix == ".xml":
        return import_xml(p)
    if suffix == ".docx":
        return import_docx(p)
    return import_txt(p)


def dedup_records(records: Iterable[SearchRecord]) -> list[SearchRecord]:
    seen = set()
    out = []
    for r in records:
        key = (r.link or "").strip()
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def import_csv(path: Path) -> list[SearchRecord]:
    records = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader, 1):
            rec = make_record(row, i)
            if rec:
                records.append(rec)
    return dedup_records(records)


def import_xlsx(path: Path) -> list[SearchRecord]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(x or "").strip() for x in rows[0]]
    records = []
    for i, values in enumerate(rows[1:], 1):
        row = {headers[j]: values[j] if j < len(values) else "" for j in range(len(headers))}
        rec = make_record(row, i)
        if rec:
            records.append(rec)
    return dedup_records(records)


def import_xml(path: Path) -> list[SearchRecord]:
    records = []
    root = ET.parse(path).getroot()
    for i, elem in enumerate(root.findall(".//record"), 1):
        row = {child.tag: child.text or "" for child in list(elem)}
        rec = make_record(row, i)
        if rec:
            records.append(rec)
    if records:
        return dedup_records(records)
    # Fallback for generic XML with embedded URLs.
    return import_urls_from_text(path.read_text(encoding="utf-8", errors="ignore"))


def import_docx(path: Path) -> list[SearchRecord]:
    from docx import Document
    doc = Document(path)
    text = "\n".join(p.text for p in doc.paragraphs)
    return import_urls_from_text(text)


def import_txt(path: Path) -> list[SearchRecord]:
    text = path.read_text(encoding="utf-8-sig", errors="ignore")
    records = []
    current_title = ""
    current_source = ""
    current_time = ""
    rank = 0
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        m_title = re.match(r"^\[(\d+)\]\s*(.*)$", s)
        if m_title:
            current_title = m_title.group(2).strip()
            try:
                rank = int(m_title.group(1))
            except Exception:
                rank += 1
            continue
        if s.lower().startswith("time:"):
            # Export format: Time: ... | Source: ...
            parts = [p.strip() for p in s.split("|")]
            if parts:
                current_time = parts[0].split(":", 1)[-1].strip()
            for part in parts:
                if part.lower().startswith("source:"):
                    current_source = part.split(":", 1)[-1].strip()
            continue
        if s.lower().startswith("url:"):
            url = s.split(":", 1)[-1].strip()
            rec = make_record({"link": url, "title": current_title or url, "source": current_source, "published_time": current_time, "rank": rank or len(records) + 1}, len(records) + 1)
            if rec:
                records.append(rec)
            continue
    if records:
        return dedup_records(records)
    return import_urls_from_text(text)


def import_urls_from_text(text: str) -> list[SearchRecord]:
    records = []
    for i, url in enumerate(URL_RE.findall(text or ""), 1):
        rec = make_record({"link": url.strip(), "title": url.strip(), "rank": i}, i)
        if rec:
            records.append(rec)
    return dedup_records(records)
