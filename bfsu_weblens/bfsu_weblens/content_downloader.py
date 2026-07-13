# -*- coding: utf-8 -*-
"""Content download, extraction, cleaning, and file-cache utilities for BFSU WebLens.

This module intentionally does not use a database. It writes raw HTML and
subsequent derived files under a user-selected content folder and appends one
JSONL manifest record per processed URL. It also maintains a human-readable
metadata workbook for corpus-management workflows.
"""
from __future__ import annotations

from dataclasses import dataclass, asdict
from datetime import datetime
import hashlib
import html
import json
import random
import re
import threading
import time
import traceback
from pathlib import Path
from typing import Any, Optional
from urllib.parse import urlparse, urlunparse, parse_qsl, urlencode

import requests
from bs4 import BeautifulSoup

TRACKING_KEYS = {
    "utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content",
    "utm_name", "utm_cid", "utm_reader", "utm_viz_id", "utm_pubreferrer",
    "fbclid", "gclid", "dclid", "mc_cid", "mc_eid", "igshid", "ref", "ref_src",
}

BASE_BOILERPLATE_PATTERNS = [
    r"^\s*(advertisement|ad feedback|sponsored content)\s*$",
    r"^\s*(sign in|log in|subscribe|subscribe now|newsletter sign[- ]?up)\s*$",
    r"^\s*(share this|share this article|follow us|follow .* on)\s*$",
    r"^\s*(cookie policy|privacy policy|terms of use|all rights reserved)\s*$",
    r"^\s*(read more|more on this story|related articles?)\s*$",
    # Common Chinese news/app shell noise.  These are deliberately conservative:
    # short UI-only lines are removed, while normal article paragraphs containing
    # the same words are kept.
    r"^\s*(登录|注册|分享|收藏|字号|大|中|小|打印|扫一扫|下载客户端|打开客户端)\s*$",
    r"^\s*(相关推荐|热门推荐|相关新闻|更多新闻|相关阅读|我要举报|特别声明)\s*$",
]

NEWS_BOILERPLATE_PATTERNS = [
    r"^\s*(listen to this article|save article|print article|email article)\s*$",
    r"^\s*(by .*?\s*)?updated \d+.*$",
    r"^\s*(our standards|the trust principles|correction:).*$",
    r"^\s*(image source|image caption|getty images|reuters|associated press)\s*$",
]

STRICT_BOILERPLATE_PATTERNS = [
    r"^\s*(cookies?|privacy|terms|copyright|©|all rights reserved).*$",
    r"^\s*(sign up|sign in|log in|register|subscribe|newsletter|follow|share|comment).*$",
    r"^\s*(more from|related|recommended|popular|trending|latest|read next).*$",
    r"^\s*(home|world|business|politics|sports|technology|culture|opinion)\s*$",
]

USER_AGENT_FALLBACK = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/152.0.0.0 Safari/537.36"
)

METADATA_EXCEL_FIELDS = [
    "processed_at", "ok", "url", "final_url", "canonical_url", "domain",
    "fetch_method", "selenium_fallback_used", "status_code", "content_type", "encoding",
    "title", "author", "byline", "publisher", "publication", "organization",
    "site_name", "source_name", "section", "keywords", "dateline", "location",
    "published_time", "modified_time", "language", "description", "article_id",
    "copyright_holder", "copyright_year", "is_accessible_for_free", "image_url",
    "publisher_logo", "search_engine", "search_query", "query_raw", "search_vertical",
    "source_filter", "sort_mode", "site_limit", "actual_domain", "date_filter_type",
    "date_start", "date_end", "start_ts", "end_ts", "baidu_gpc", "search_url",
    "source", "rank",
    "extraction_method", "cleaning_scheme", "char_count", "word_count", "paragraph_count",
    "quality_score", "raw_html_path", "raw_bytes_path", "raw_text_path", "clean_text_path",
    "metadata_path", "request_error", "selenium_error", "error",
]


@dataclass
class ContentDownloadSettings:
    content_root: str
    max_workers: int = 3
    timeout_seconds: int = 30
    user_agent: str = USER_AGENT_FALLBACK
    min_delay_ms: int = 0
    max_delay_ms: int = 0
    # requests | selenium | mixed. Mixed means requests first, Selenium fallback when needed.
    fetch_mode: str = "mixed"
    receive_wait_ms: int = 0
    cleaning_scheme: str = "auto"  # auto | newspaper_news | news_article | general_web | strict_corpus | light_archive
    selenium_fallback: bool = True
    selenium_backend: str = "selenium_chrome"  # selenium_chrome | selenium_edge
    selenium_driver_path: str = ""
    selenium_binary_path: str = ""
    selenium_wait_ms: int = 3500
    selenium_headless: bool = False
    retry_count: int = 1
    task_timeout_seconds: int = 300
    resume_enabled: bool = True
    domain_lock_timeout_seconds: int = 300


@dataclass
class ContentResult:
    ok: bool
    url: str
    final_url: str = ""
    domain: str = ""
    status_code: int = 0
    content_type: str = ""
    encoding: str = ""
    title: str = ""
    author: str = ""
    published_time: str = ""
    language: str = ""
    extraction_method: str = ""
    cleaning_scheme: str = ""
    char_count: int = 0
    word_count: int = 0
    quality_score: float = 0.0
    raw_html_path: str = ""
    raw_text_path: str = ""
    clean_text_path: str = ""
    metadata_path: str = ""
    metadata_excel_path: str = ""
    manifest_path: str = ""
    fetch_method: str = ""
    selenium_fallback_used: bool = False
    error: str = ""


class DomainLockPool:
    """A lock registry that ensures only one worker hits the same domain at once."""

    def __init__(self) -> None:
        self._locks: dict[str, threading.Lock] = {}
        self._guard = threading.Lock()

    def lock_for(self, domain: str) -> threading.Lock:
        key = (domain or "unknown").lower()
        with self._guard:
            if key not in self._locks:
                self._locks[key] = threading.Lock()
            return self._locks[key]


def utc_now_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def normalize_url(url: str) -> str:
    """Normalize a URL for content-cache identity without removing meaningful query keys."""
    url = (url or "").strip()
    if not url:
        return ""
    try:
        parsed = urlparse(url)
        scheme = (parsed.scheme or "https").lower()
        netloc = parsed.netloc.lower()
        path = parsed.path or "/"
        if path != "/":
            path = path.rstrip("/")
        query_items = []
        for k, v in parse_qsl(parsed.query, keep_blank_values=True):
            kl = k.lower()
            if kl in TRACKING_KEYS or kl.startswith("utm_"):
                continue
            query_items.append((k, v))
        query = urlencode(sorted(query_items), doseq=True)
        return urlunparse((scheme, netloc, path, "", query, ""))
    except Exception:
        return url


def domain_of(url: str) -> str:
    try:
        return urlparse(url).netloc.lower().lstrip("www.")
    except Exception:
        return ""


def safe_slug(text: str, max_len: int = 64) -> str:
    text = re.sub(r"\s+", "_", (text or "").strip())
    text = re.sub(r"[^\w\-.\u4e00-\u9fff]+", "_", text, flags=re.UNICODE)
    text = text.strip("._-")
    return (text[:max_len] or "untitled")


def record_url(record: Any) -> str:
    if isinstance(record, dict):
        return str(record.get("link") or record.get("url") or record.get("final_url") or "")
    return str(getattr(record, "link", "") or getattr(record, "url", "") or "")


def record_title(record: Any) -> str:
    if isinstance(record, dict):
        return str(record.get("title") or "")
    return str(getattr(record, "title", "") or "")


def record_dict(record: Any) -> dict[str, Any]:
    if hasattr(record, "to_dict"):
        return record.to_dict()
    if isinstance(record, dict):
        return dict(record)
    return {}


def ensure_content_dirs(root: str | Path) -> dict[str, Path]:
    root_path = Path(root).expanduser().resolve()
    paths = {
        "root": root_path,
        "raw_html": root_path / "raw_html",
        "raw_text": root_path / "raw_text",
        "clean_text": root_path / "clean_text",
        "metadata": root_path / "metadata",
        "logs": root_path / "logs",
        "exports": root_path / "exports",
    }
    for p in paths.values():
        p.mkdir(parents=True, exist_ok=True)
    return paths


def content_url_key(url: str) -> str:
    return normalize_url(url or "")


def _path_exists(path_text: str) -> bool:
    if not path_text:
        return False
    try:
        return Path(path_text).exists()
    except Exception:
        return False


def load_successful_download_index(content_root: str | Path) -> dict[str, dict[str, Any]]:
    """Read content_manifest.jsonl and return the latest successful item per URL.

    This is the download-only checkpoint mechanism.  It is deliberately scoped to
    content downloads and is never used for Google/Baidu search crawling, so a
    forced program close will not resume search-engine pagination automatically.
    """
    root = Path(content_root).expanduser().resolve()
    manifest = root / "content_manifest.jsonl"
    index: dict[str, dict[str, Any]] = {}
    if not manifest.exists():
        return index
    try:
        with manifest.open("r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    item = json.loads(line)
                except Exception:
                    continue
                if not item.get("ok"):
                    continue
                paths = item.get("paths") or {}
                # A successful checkpoint must have at least one persisted text/HTML artifact.
                if not any(_path_exists(str(paths.get(k) or "")) for k in ("clean_text", "raw_text", "raw_html")):
                    continue
                for key_url in (item.get("url"), item.get("final_url"), item.get("canonical_url")):
                    key = content_url_key(str(key_url or ""))
                    if key:
                        index[key] = item
    except Exception:
        return index
    return index


def successful_manifest_for_record(record: Any, success_index: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    candidates = [record_url(record)]
    data = record_dict(record)
    for k in ("url", "link", "final_url", "canonical_url"):
        v = data.get(k)
        if v:
            candidates.append(str(v))
    for u in candidates:
        key = content_url_key(u)
        if key and key in success_index:
            return success_index[key]
    return None


def content_result_from_manifest(record: Any, metadata: dict[str, Any]) -> ContentResult:
    paths = metadata.get("paths") or {}
    meta = metadata.get("metadata") or {}
    counts = metadata.get("counts") or {}
    return ContentResult(
        ok=True,
        url=record_url(record) or str(metadata.get("url") or ""),
        final_url=str(metadata.get("final_url") or metadata.get("url") or record_url(record) or ""),
        domain=str(metadata.get("domain") or domain_of(str(metadata.get("final_url") or metadata.get("url") or record_url(record) or ""))),
        status_code=int(metadata.get("status_code") or 0),
        content_type=str(metadata.get("content_type") or ""),
        encoding=str(metadata.get("encoding") or ""),
        title=str(meta.get("title") or record_title(record) or ""),
        author=str(meta.get("author") or ""),
        published_time=str(meta.get("published_time") or ""),
        language=str(meta.get("language") or ""),
        extraction_method="resume_checkpoint",
        cleaning_scheme=str(metadata.get("cleaning_scheme") or ""),
        char_count=int((counts.get("chars") if isinstance(counts, dict) else 0) or 0),
        word_count=int((counts.get("words") if isinstance(counts, dict) else 0) or 0),
        quality_score=float(metadata.get("quality_score") or 0),
        raw_html_path=str(paths.get("raw_html") or ""),
        raw_text_path=str(paths.get("raw_text") or ""),
        clean_text_path=str(paths.get("clean_text") or ""),
        metadata_path=str(paths.get("metadata") or ""),
        metadata_excel_path=str(Path(metadata.get("manifest_path", "")).with_name("content_metadata.xlsx") if metadata.get("manifest_path") else ""),
        manifest_path=str(metadata.get("manifest_path") or ""),
        fetch_method=str(metadata.get("fetch_method") or ""),
        selenium_fallback_used=bool(metadata.get("selenium_fallback_used")),
    )


def cache_stem(record: Any, index: int = 0) -> str:
    url = normalize_url(record_url(record))
    h = hashlib.sha1(url.encode("utf-8", errors="ignore")).hexdigest()[:12]
    dom = safe_slug(domain_of(url), 40)
    title = safe_slug(record_title(record), 48)
    prefix = f"{index:05d}_" if index else ""
    return f"{prefix}{dom}_{title}_{h}"


def _as_clean_text(value: Any) -> str:
    """Convert common metadata values into a readable single-line string."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float, bool)):
        return str(value)
    if isinstance(value, dict):
        for key in ("name", "headline", "text", "url", "@id", "identifier"):
            if value.get(key):
                return _as_clean_text(value.get(key))
        return ""
    if isinstance(value, list):
        parts = [_as_clean_text(v) for v in value]
        return "; ".join([p for p in parts if p])
    return str(value).strip()


def _extract_jsonld_location(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, list):
        return "; ".join([_extract_jsonld_location(v) for v in value if _extract_jsonld_location(v)])
    if isinstance(value, dict):
        parts: list[str] = []
        for key in ("name", "addressLocality", "addressRegion", "addressCountry"):
            txt = _as_clean_text(value.get(key))
            if txt and txt not in parts:
                parts.append(txt)
        addr = value.get("address")
        if addr:
            txt = _extract_jsonld_location(addr)
            if txt and txt not in parts:
                parts.append(txt)
        return ", ".join(parts)
    return str(value).strip()


def _set_meta(meta: dict[str, str], key: str, value: Any, overwrite: bool = False) -> None:
    text = _as_clean_text(value)
    text = re.sub(r"\s+", " ", text or "").strip()
    if text and (overwrite or not meta.get(key)):
        meta[key] = text


def _set_location_meta(meta: dict[str, str], key: str, value: Any, overwrite: bool = False) -> None:
    text = _extract_jsonld_location(value)
    text = re.sub(r"\s+", " ", text or "").strip()
    if text and (overwrite or not meta.get(key)):
        meta[key] = text


def extract_json_ld_meta(soup: BeautifulSoup) -> dict[str, str]:
    """Extract corpus-relevant metadata from schema.org JSON-LD.

    News sites vary substantially.  This function therefore accepts Article,
    NewsArticle, BlogPosting, WebPage and CreativeWork records and keeps source
    evidence such as publisher, organization, section, keywords, dateline and
    location whenever available.
    """
    meta: dict[str, str] = {}
    scripts = list(soup.find_all("script", attrs={"type": re.compile(r"application/ld\+json", re.I)}))
    for tag in scripts:
        text = tag.string or tag.get_text(" ", strip=True)
        if not text:
            continue
        try:
            data = json.loads(text)
        except Exception:
            # Some sites wrap JSON-LD arrays with harmless whitespace/comments; keep conservative.
            continue
        items = data if isinstance(data, list) else [data]
        idx = 0
        while idx < len(items):
            item = items[idx]
            idx += 1
            if not isinstance(item, dict):
                continue
            graph = item.get("@graph")
            if isinstance(graph, list):
                items.extend([x for x in graph if isinstance(x, dict)])
            typ = item.get("@type")
            typ_text = " ".join(typ) if isinstance(typ, list) else str(typ or "")
            if typ_text and not re.search(r"Article|NewsArticle|BlogPosting|ReportageNewsArticle|WebPage|CreativeWork", typ_text, re.I):
                # Organization nodes inside @graph are still useful as publisher evidence.
                if re.search(r"Organization|NewsMediaOrganization|LocalBusiness", typ_text, re.I):
                    _set_meta(meta, "organization", item.get("name"))
                    _set_meta(meta, "publisher", item.get("name"))
                continue
            _set_meta(meta, "schema_type", typ_text)
            _set_meta(meta, "title", item.get("headline") or item.get("name"))
            _set_meta(meta, "description", item.get("description"))
            _set_meta(meta, "published_time", item.get("datePublished") or item.get("dateCreated"))
            _set_meta(meta, "modified_time", item.get("dateModified"))
            _set_meta(meta, "article_id", item.get("identifier") or item.get("@id"))
            _set_meta(meta, "url", item.get("url") or item.get("mainEntityOfPage"))
            _set_meta(meta, "author", item.get("author") or item.get("creator"))
            _set_meta(meta, "byline", item.get("byline"))
            _set_meta(meta, "publisher", item.get("publisher"))
            _set_meta(meta, "publication", item.get("isPartOf") or item.get("sourceOrganization"))
            _set_meta(meta, "organization", item.get("sourceOrganization") or item.get("provider"))
            _set_meta(meta, "section", item.get("articleSection") or item.get("genre"))
            _set_meta(meta, "keywords", item.get("keywords"))
            _set_meta(meta, "dateline", item.get("dateline"))
            _set_location_meta(meta, "location", item.get("contentLocation") or item.get("locationCreated") or item.get("spatialCoverage"))
            _set_meta(meta, "language", item.get("inLanguage"))
            _set_meta(meta, "copyright_holder", item.get("copyrightHolder"))
            _set_meta(meta, "copyright_year", item.get("copyrightYear"))
            _set_meta(meta, "is_accessible_for_free", item.get("isAccessibleForFree"))
            image = item.get("image") or item.get("thumbnailUrl")
            _set_meta(meta, "image_url", image)
            publisher = item.get("publisher")
            if isinstance(publisher, dict):
                logo = publisher.get("logo")
                if isinstance(logo, dict):
                    _set_meta(meta, "publisher_logo", logo.get("url"))
    return {k: v for k, v in meta.items() if v}

def get_meta_content(soup: BeautifulSoup, *names: str) -> str:
    for name in names:
        tag = soup.find("meta", attrs={"property": name}) or soup.find("meta", attrs={"name": name})
        if tag and tag.get("content"):
            return html.unescape(str(tag.get("content", "")).strip())
    return ""


# ---------------------------------------------------------------------------
# Multilingual decoding and Chinese/mojibake repair helpers
# ---------------------------------------------------------------------------
MOJIBAKE_PATTERN = re.compile(
    r"(Ã.|Â.|â[€œ\x80-\xbf]|ä[¸-¿]|å[\x80-¿]|æ[\x80-¿]|ç[\x80-¿]|è[\x80-¿]|é[\x80-¿]|ï¼|ã|ðŸ)",
    re.I,
)
REPLACEMENT_CHARS = {"\ufffd", "�"}


def _script_readability_score(text: str) -> float:
    """Score decoded text by readable script coverage and mojibake penalties."""
    if not text:
        return -999999.0
    sample = text[:200000]
    length = max(1, len(sample))
    replacement = sum(sample.count(ch) for ch in REPLACEMENT_CHARS)
    mojibake_hits = len(MOJIBAKE_PATTERN.findall(sample))
    readable = len(re.findall(
        r"[A-Za-z0-9\u4e00-\u9fff\u3040-\u30ff\u3400-\u4dbf\uac00-\ud7af\u0400-\u04ff\u0370-\u03ff\u0600-\u06ff\u0900-\u097f\u0980-\u09ff\u0e00-\u0e7f\u00c0-\u024f]",
        sample,
    ))
    cjk = len(re.findall(r"[\u4e00-\u9fff]", sample))
    whitespace = len(re.findall(r"\s", sample))
    # Text with almost no spaces can be perfectly valid CJK, so reward CJK too.
    return (readable / length) * 100.0 + min(cjk, 2000) * 0.02 - mojibake_hits * 6.0 - replacement * 25.0 - max(0, whitespace - length * 0.55) * 0.01


def looks_mojibake(text: str) -> bool:
    if not text:
        return False
    sample = text[:80000]
    if sum(sample.count(ch) for ch in REPLACEMENT_CHARS) >= 3:
        return True
    hits = len(MOJIBAKE_PATTERN.findall(sample))
    # Typical UTF-8-as-Latin1 Chinese appears as many ä/å/æ/é sequences.
    return hits >= 4 or (hits >= 2 and len(sample) < 300)


def repair_mojibake_text(text: str) -> str:
    """Repair common UTF-8 decoded as Latin-1/CP1252 without touching normal text.

    This is essential for many Chinese news sites returned by Baidu, but it is
    intentionally generic: it also fixes Western smart-quote mojibake while
    leaving valid English, Spanish, Arabic, Japanese, Korean, etc. unchanged.
    """
    if not text or not looks_mojibake(text):
        return text or ""
    candidates = [text]
    for enc in ("latin1", "cp1252"):
        try:
            candidates.append(text.encode(enc, errors="strict").decode("utf-8", errors="strict"))
        except Exception:
            pass
    try:
        candidates.append(text.encode("latin1", errors="ignore").decode("utf-8", errors="ignore"))
    except Exception:
        pass
    best = max(candidates, key=_script_readability_score)
    # Use the repaired version only when it is materially better.
    if _script_readability_score(best) > _script_readability_score(text) + 10:
        return best
    return text


def _charset_from_content_type(content_type: str) -> str:
    m = re.search(r"charset\s*=\s*['\"]?([^;'\"\s>]+)", content_type or "", re.I)
    return (m.group(1).strip() if m else "")


def _charset_from_meta(raw: bytes) -> str:
    head = raw[:20000].decode("latin1", errors="ignore")
    patterns = [
        r"<meta[^>]+charset=['\"]?([^'\"\s/>]+)",
        r"<meta[^>]+content=['\"][^'\"]*charset=([^'\"\s;>]+)",
    ]
    for pat in patterns:
        m = re.search(pat, head, re.I)
        if m:
            return m.group(1).strip()
    return ""


def decode_html_bytes(raw: bytes, content_type: str = "", apparent_encoding: str = "", declared_encoding: str = "") -> tuple[str, str, str]:
    """Decode raw HTML bytes robustly for multilingual pages.

    The function considers HTTP headers, meta charset, requests' apparent
    encoding, charset-normalizer when available, and common Asian encodings. It
    then applies mojibake repair only if the repaired text scores better.
    """
    raw = raw or b""
    if not raw:
        return "", "", "empty"
    candidates: list[str] = []
    # BOMs first.
    if raw.startswith(b"\xef\xbb\xbf"):
        candidates.append("utf-8-sig")
    elif raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff"):
        candidates.append("utf-16")
    header_enc = _charset_from_content_type(content_type)
    meta_enc = _charset_from_meta(raw)
    for enc in [declared_encoding, header_enc, meta_enc, apparent_encoding]:
        if enc and enc.lower() not in {"none", "unknown", "binary"}:
            candidates.append(enc)
    try:
        from charset_normalizer import from_bytes  # type: ignore
        best = from_bytes(raw).best()
        if best and getattr(best, "encoding", None):
            candidates.append(str(best.encoding))
    except Exception:
        pass
    # Multilingual fallbacks.  gb18030 covers GBK/GB2312; cp932 helps Japanese;
    # euc-kr helps Korean; windows-1252/latin1 is last-resort Western fallback.
    candidates.extend(["utf-8", "gb18030", "big5", "cp950", "shift_jis", "cp932", "euc-kr", "windows-1252", "latin1"])
    seen = set()
    decoded: list[tuple[str, str]] = []
    for enc in candidates:
        enc_norm = (enc or "").strip().lower().replace("_", "-")
        if not enc_norm or enc_norm in seen:
            continue
        seen.add(enc_norm)
        try:
            txt = raw.decode(enc, errors="replace")
        except Exception:
            continue
        txt = repair_mojibake_text(txt)
        decoded.append((txt, enc))
    if not decoded:
        txt = raw.decode("utf-8", errors="replace")
        return repair_mojibake_text(txt), "utf-8", "fallback_replace"
    best_txt, best_enc = max(decoded, key=lambda item: _script_readability_score(item[0]))
    return best_txt, best_enc, "auto"


def decode_response_html(response: requests.Response) -> tuple[str, str, str]:
    raw = getattr(response, "content", b"") or b""
    content_type = response.headers.get("Content-Type", "") if response is not None else ""
    apparent = getattr(response, "apparent_encoding", "") or ""
    declared = getattr(response, "encoding", "") or ""
    return decode_html_bytes(raw, content_type, apparent, declared)


# ---------------------------------------------------------------------------
# Site-specific and generic article extraction helpers
# ---------------------------------------------------------------------------
SITE_ARTICLE_SELECTORS: list[tuple[str, list[str]]] = [
    (r"(10jqka\.com\.cn|ths\.cn)", [".news-content-parsed", ".news-content.article-content", ".article-content"]),
    (r"bbtnews\.com\.cn", ["#pageContent", ".article-bd", ".article-box"]),
    (r"myzaker\.com|app\.myzaker\.com", ["#content", "#content_text", ".article_content"]),
    (r"jjckb\.cn", ["#detailContent", "#content", "#detail .mainCon", "#detail"]),
    (r"thepaper\.cn", ["[class*=cententWrap]", "[class*=contentWrap]", "[class*=normalContentWrap] article", "article"]),
    (r"eastmoney\.com", ["#ContentBody", ".txtinfos", ".contentbox .mainleft"]),
]

SITE_META_SELECTORS: list[tuple[str, dict[str, list[str]]]] = [
    (r"myzaker\.com|app\.myzaker\.com", {"title": ["#tpl_title"], "author": ["#tpl_author"], "published_time": ["#tpl_date"]}),
    (r"bbtnews\.com\.cn", {"title": [".article-hd h3"], "source_name": [".assist .info span"], "published_time": [".assist .info span"]}),
    (r"jjckb\.cn", {"title": [".top_tit"], "source_name": [".sj_scro span"], "published_time": [".sj_scro span"]}),
    (r"thepaper\.cn", {"title": ["h1", "[class*=title]"], "published_time": ["[class*=headerContent] span"], "source_name": ["[class*=headerContent] span"]}),
    (r"eastmoney\.com", {"title": ["#topbox .title"], "published_time": ["#topbox .infos .item"], "source_name": ["#topbox .infos .item", ".sourcebox span"]}),
]


def _domain_matches(url: str, pattern: str) -> bool:
    return re.search(pattern, domain_of(url) or url, re.I) is not None


def _texts_by_selectors(soup: BeautifulSoup, selectors: list[str]) -> list[str]:
    texts: list[str] = []
    for sel in selectors:
        try:
            for tag in soup.select(sel):
                text = tag.get_text(" ", strip=True)
                text = repair_mojibake_text(normalize_unicode_and_space(text))
                if text:
                    texts.append(text)
        except Exception:
            continue
    return texts


def _first_text_by_selectors(soup: BeautifulSoup, selectors: list[str]) -> str:
    texts = _texts_by_selectors(soup, selectors)
    return texts[0] if texts else ""


def extract_site_selector_meta(soup: BeautifulSoup, url: str) -> dict[str, str]:
    meta: dict[str, str] = {}
    for pattern, selectors_by_key in SITE_META_SELECTORS:
        if not _domain_matches(url, pattern):
            continue
        for key, selectors in selectors_by_key.items():
            texts = _texts_by_selectors(soup, selectors)
            value = ""
            if key == "published_time":
                for txt in texts:
                    m = re.search(r"\d{4}[-年/]\d{1,2}[-月/]\d{1,2}(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?", txt)
                    if m:
                        value = m.group(0)
                        break
                # Allow relative Chinese dates only as last resort.
                if not value:
                    for txt in texts:
                        if re.fullmatch(r"(今天|昨天|前天|\d+分钟前|\d+小时前|\d+天前)", txt):
                            value = txt
                            break
            elif key == "source_name":
                for txt in texts:
                    m = re.search(r"(?:来源|出处|文章来源)[:：]\s*([^\s　|；;]+)", txt)
                    if m:
                        value = m.group(1)
                        break
                if not value:
                    for txt in texts:
                        # Prefer short non-date source labels.
                        if not re.search(r"\d{4}[-年/]\d{1,2}", txt) and 1 <= len(txt) <= 30:
                            value = re.sub(r"^(来源|出处|文章来源)[:：]\s*", "", txt).strip()
                            break
            else:
                for txt in texts:
                    txt2 = re.sub(r"^(来源|出处|作者|网编|时间|日期)[:：]\s*", "", txt).strip()
                    if txt2:
                        value = txt2
                        break
            if value and not meta.get(key):
                meta[key] = value
    return meta



def candidate_text_score(candidate: Any) -> tuple[float, str]:
    text = block_text_from_candidate(candidate, include_lists=False, min_chars=2)
    text = normalize_unicode_and_space(text)
    if not text:
        return -999999.0, ""
    paras = [p for p in re.split(r"\n\s*\n", text) if p.strip()]
    link_text = " ".join(a.get_text(" ", strip=True) for a in candidate.find_all("a")) if hasattr(candidate, "find_all") else ""
    link_ratio = len(link_text) / max(1, len(text))
    cls_id = ""
    if hasattr(candidate, "get"):
        try:
            attrs = getattr(candidate, "attrs", {}) or {}
            cls_val = attrs.get("class") or []
            if isinstance(cls_val, str):
                cls_val = [cls_val]
            cls_id = " ".join([str(attrs.get("id") or ""), " ".join(str(x) for x in cls_val)]).lower()
        except Exception:
            cls_id = ""
    bonus = 0
    if re.search(r"article|content|正文|detail|story|post|main|entry|txt|text|news", cls_id):
        bonus += 400
    if getattr(candidate, "name", "") in {"article", "main"}:
        bonus += 300
    penalty = 0
    if re.search(r"nav|menu|footer|header|comment|share|related|recommend|ad|advert|sidebar|login|download", cls_id):
        penalty += 900
    return len(text) + len(paras) * 120 + bonus - link_ratio * 1200 - penalty, text


def generic_article_candidate_text(soup: BeautifulSoup) -> tuple[str, str]:
    candidates: list[Any] = []
    for sel in ["article", "main", "[role=main]", ".article", ".content", ".post", ".entry", ".story", ".main", "#article", "#content", "#main", "#detail"]:
        try:
            candidates.extend(soup.select(sel))
        except Exception:
            pass
    rx = re.compile(r"article|content|正文|detail|story|post|main|entry|txt|text|news", re.I)
    for tag in soup.find_all(True):
        try:
            attrs = getattr(tag, "attrs", {}) or {}
            cls_val = attrs.get("class") or []
            if isinstance(cls_val, str):
                cls_val = [cls_val]
            val = " ".join([str(attrs.get("id") or ""), " ".join(str(x) for x in cls_val)])
            if val and rx.search(val):
                candidates.append(tag)
        except Exception:
            continue
    if not candidates:
        candidates = [soup.find("article") or soup.find("main") or soup.body or soup]
    unique = []
    seen_ids = set()
    for c in candidates:
        if id(c) not in seen_ids:
            unique.append(c)
            seen_ids.add(id(c))
    scored = [candidate_text_score(c) for c in unique]
    scored = [x for x in scored if x[1]]
    if not scored:
        return "", "bs4_visible_text"
    best_score, best_text = max(scored, key=lambda x: x[0])
    return best_text, "bs4_best_candidate"


def extract_with_site_templates(html_text: str, url: str) -> tuple[str, str]:
    soup = prepare_soup(html_text, remove_layout=True)
    for pattern, selectors in SITE_ARTICLE_SELECTORS:
        if not _domain_matches(url, pattern):
            continue
        for sel in selectors:
            try:
                for tag in soup.select(sel):
                    text = block_text_from_candidate(tag, include_lists=False, min_chars=1)
                    text = repair_mojibake_text(normalize_unicode_and_space(text))
                    if is_usable_text(text, min_chars=80):
                        return text, f"site_template:{domain_of(url)}:{sel}"
            except Exception:
                continue
    text, method = generic_article_candidate_text(soup)
    if text:
        return text, method
    return extract_bs4_article_text(html_text, remove_layout=True), "bs4_news_article_fallback"


def extract_metadata(html_text: str, url: str) -> dict[str, str]:
    soup = BeautifulSoup(html_text or "", "html.parser")
    meta = extract_json_ld_meta(soup)
    h1 = soup.find("h1")
    title_tag = soup.find("title")
    html_tag = soup.find("html")

    meta.setdefault("title", get_meta_content(soup, "og:title", "twitter:title", "title", "parsely-title", "sailthru.title") or (h1.get_text(" ", strip=True) if h1 else "") or (title_tag.get_text(" ", strip=True) if title_tag else ""))
    meta.setdefault("description", get_meta_content(soup, "og:description", "twitter:description", "description", "dc.description", "DC.description"))
    meta.setdefault("published_time", get_meta_content(
        soup,
        "article:published_time", "pubdate", "publishdate", "date", "datePublished",
        "dc.date", "DC.date", "DC.date.issued", "citation_publication_date",
        "sailthru.date", "parsely-pub-date", "ptime",
    ))
    meta.setdefault("modified_time", get_meta_content(soup, "article:modified_time", "lastmod", "modified", "dateModified", "dc.date.modified", "DC.date.modified"))
    meta.setdefault("author", get_meta_content(
        soup,
        "article:author", "author", "byl", "byline", "dc.creator", "DC.creator",
        "citation_author", "sailthru.author", "parsely-author", "twitter:creator",
    ))
    meta.setdefault("byline", get_meta_content(soup, "byline", "parsely-author", "sailthru.author"))
    meta.setdefault("publisher", get_meta_content(soup, "publisher", "dc.publisher", "DC.publisher", "citation_publisher", "article:publisher"))
    meta.setdefault("publication", get_meta_content(soup, "citation_journal_title", "citation_conference_title", "prism.publicationName", "sailthru.source"))
    meta.setdefault("site_name", get_meta_content(soup, "og:site_name", "application-name", "twitter:site"))
    meta.setdefault("source_name", get_meta_content(soup, "og:site_name", "sailthru.source", "parsely-network-canonical"))
    meta.setdefault("section", get_meta_content(soup, "article:section", "section", "parsely-section", "sailthru.vertical", "prism.section"))
    meta.setdefault("keywords", get_meta_content(soup, "news_keywords", "keywords", "article:tag", "sailthru.tags", "parsely-tags", "citation_keywords"))
    meta.setdefault("dateline", get_meta_content(soup, "dateline", "article:dateline", "parsely-post-id"))
    meta.setdefault("location", get_meta_content(soup, "geo.placename", "geo.position", "ICBM", "dc.coverage", "DC.coverage", "location"))
    meta.setdefault("article_id", get_meta_content(soup, "article:id", "parsely-post-id", "sailthru.post_id", "citation_doi", "dc.identifier", "DC.identifier"))
    meta.setdefault("copyright_holder", get_meta_content(soup, "copyright", "dc.rights", "DC.rights", "rights"))
    meta.setdefault("copyright_year", get_meta_content(soup, "copyrightYear"))
    meta.setdefault("is_accessible_for_free", get_meta_content(soup, "isAccessibleForFree"))
    meta.setdefault("image_url", get_meta_content(soup, "og:image", "twitter:image"))
    meta.setdefault("publisher_logo", get_meta_content(soup, "logo"))
    meta.setdefault("canonical_url", "")
    canonical = soup.find("link", attrs={"rel": re.compile(r"canonical", re.I)})
    if canonical and canonical.get("href"):
        meta["canonical_url"] = str(canonical.get("href", "")).strip()
    if html_tag and html_tag.get("lang"):
        meta.setdefault("language", str(html_tag.get("lang", "")).strip())
    meta.setdefault("language", get_meta_content(soup, "og:locale", "language", "dc.language", "DC.language"))
    meta.setdefault("domain", domain_of(url))
    meta.setdefault("url", url)
    selector_meta = extract_site_selector_meta(soup, url)
    for _k, _v in selector_meta.items():
        if _v and not meta.get(_k):
            meta[_k] = _v
    if meta.get("site_name") and not meta.get("publisher"):
        meta["publisher"] = meta["site_name"]
    if meta.get("publisher") and not meta.get("organization"):
        meta["organization"] = meta["publisher"]
    if meta.get("site_name") and not meta.get("source_name"):
        meta["source_name"] = meta["site_name"]
    cleaned_meta: dict[str, str] = {}
    for k, v in meta.items():
        if v is None or not str(v).strip():
            continue
        txt = repair_mojibake_text(str(v))
        txt = re.sub(r"\s+", " ", txt).strip()
        if txt:
            cleaned_meta[k] = txt
    return cleaned_meta

def prepare_soup(html_text: str, remove_layout: bool = False) -> BeautifulSoup:
    soup = BeautifulSoup(html_text or "", "html.parser")
    remove = ["script", "style", "noscript", "svg", "canvas", "form", "iframe"]
    if remove_layout:
        remove += ["nav", "header", "footer", "aside"]
    for tag in soup(remove):
        tag.decompose()
    # Remove common comments/advertising containers without relying on English only.
    for tag in list(soup.find_all(True)):
        try:
            attrs = getattr(tag, "attrs", {}) or {}
            cls_val = attrs.get("class") or []
            if isinstance(cls_val, str):
                cls_val = [cls_val]
            cls_id = " ".join([str(attrs.get("id") or ""), " ".join(str(x) for x in cls_val)]).lower()
        except Exception:
            cls_id = ""
        if re.search(r"cookie|advert|ad-|ads|promo|share|social|newsletter|subscribe|comment|related|recommend", cls_id):
            tag.decompose()
    return soup


def block_text_from_candidate(candidate: Any, include_lists: bool = True, min_chars: int = 1) -> str:
    tags = ["h1", "h2", "h3", "p", "blockquote"]
    if include_lists:
        tags.append("li")
    pieces: list[str] = []
    for block in candidate.find_all(tags):
        # Skip blocks dominated by links.
        text = block.get_text(" ", strip=True)
        text = re.sub(r"\s+", " ", html.unescape(text)).strip()
        if not text or len(text) < min_chars:
            continue
        link_text = " ".join(a.get_text(" ", strip=True) for a in block.find_all("a"))
        if text and link_text and len(link_text) / max(1, len(text)) > 0.75 and len(text) < 180:
            continue
        pieces.append(text)
    if not pieces:
        text = candidate.get_text("\n", strip=True)
        pieces = [re.sub(r"\s+", " ", x).strip() for x in text.splitlines() if x.strip()]
    return "\n\n".join(pieces)


def extract_bs4_article_text(html_text: str, remove_layout: bool = True) -> str:
    soup = prepare_soup(html_text, remove_layout=remove_layout)
    text, _method = generic_article_candidate_text(soup)
    if text:
        return repair_mojibake_text(normalize_unicode_and_space(text))
    candidate = soup.find("article") or soup.find("main") or soup.body or soup
    return repair_mojibake_text(block_text_from_candidate(candidate, include_lists=True, min_chars=1))


def extract_general_web_text(html_text: str) -> str:
    soup = prepare_soup(html_text, remove_layout=True)
    text, _method = generic_article_candidate_text(soup)
    if text:
        return repair_mojibake_text(normalize_unicode_and_space(text))
    candidate = soup.find("main") or soup.find("article") or soup.body or soup
    return repair_mojibake_text(block_text_from_candidate(candidate, include_lists=True, min_chars=2))


def extract_light_archive_text(html_text: str) -> str:
    soup = prepare_soup(html_text, remove_layout=False)
    candidate = soup.find("article") or soup.find("main") or soup.body or soup
    return repair_mojibake_text(block_text_from_candidate(candidate, include_lists=True, min_chars=1))


def extract_with_newspaper(html_text: str, url: str) -> tuple[str, dict[str, str]]:
    try:
        from newspaper import Article  # type: ignore
    except Exception as exc:
        raise RuntimeError("newspaper3k is not installed or cannot be imported") from exc
    article = Article(url or "")
    article.set_html(html_text or "")
    article.parse()
    text = normalize_unicode_and_space(article.text or "")
    meta: dict[str, str] = {}
    if article.title:
        meta["title"] = article.title
    if article.authors:
        meta["author"] = "; ".join(article.authors)
    if article.publish_date:
        meta["published_time"] = article.publish_date.isoformat()
    return text, meta


def is_usable_text(text: str, min_chars: int = 180) -> bool:
    text = normalize_unicode_and_space(text)
    if len(text) < min_chars:
        return False
    paras = [p for p in re.split(r"\n\s*\n", text) if p.strip()]
    if len(paras) < 2 and len(text) < 500:
        return False
    return True


def choose_auto_scheme(record: Any, url: str) -> str:
    data = record_dict(record)
    vertical = str(data.get("search_vertical") or "").lower()
    domain = domain_of(url)
    source = str(data.get("source") or "").strip()
    if vertical == "news" or source:
        return "newspaper_news"
    if re.search(r"(\.edu|\.edu\.|\.gov|\.gov\.|\.ac\.|\.org)", domain):
        return "light_archive"
    return "general_web"


def extract_text_by_scheme(html_text: str, url: str, record: Any, scheme: str) -> tuple[str, str, str, dict[str, str]]:
    requested = (scheme or "auto").strip().lower()
    resolved = choose_auto_scheme(record, url) if requested == "auto" else requested
    extra_meta: dict[str, str] = {}
    html_text = repair_mojibake_text(html_text or "")

    def site_or_generic_news() -> tuple[str, str]:
        text, method = extract_with_site_templates(html_text, url)
        if is_usable_text(text, min_chars=100):
            return text, method
        # Last generic fallback should still return a clean visible-text body.
        return extract_bs4_article_text(html_text, remove_layout=True), "bs4_news_article_fallback"

    if resolved == "newspaper_news":
        # newspaper3k can work well on many English and some non-English news
        # pages, but it is not uniformly reliable across Chinese/Japanese/Korean,
        # app-style mobile pages, or Next.js pages.  We therefore treat it as the
        # first attempt, not as an assumed dependency.
        try:
            text, extra_meta = extract_with_newspaper(html_text, url)
            text = repair_mojibake_text(normalize_unicode_and_space(text))
            if is_usable_text(text):
                return text, "newspaper3k", resolved, extra_meta
        except Exception as exc:
            extra_meta["newspaper_error"] = str(exc)
        text, method = site_or_generic_news()
        return text, method, resolved, extra_meta

    if resolved == "news_article":
        text, method = site_or_generic_news()
        return text, method, resolved, extra_meta

    if resolved == "general_web":
        text = extract_general_web_text(html_text)
        if not is_usable_text(text, min_chars=80):
            text, method = site_or_generic_news()
            return text, method, resolved, extra_meta
        return text, "bs4_general_web", resolved, extra_meta

    if resolved == "strict_corpus":
        return extract_general_web_text(html_text), "bs4_strict_corpus", resolved, extra_meta

    if resolved == "light_archive":
        return extract_light_archive_text(html_text), "bs4_light_archive", resolved, extra_meta

    text, method = site_or_generic_news()
    return text, method, "news_article", extra_meta


def normalize_unicode_and_space(text: str) -> str:
    text = repair_mojibake_text(html.unescape(text or ""))
    text = text.replace("\u00a0", " ").replace("\ufeff", "")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def patterns_for_scheme(scheme: str) -> list[str]:
    scheme = (scheme or "").lower()
    patterns = list(BASE_BOILERPLATE_PATTERNS)
    if scheme in {"newspaper_news", "news_article", "strict_corpus"}:
        patterns += NEWS_BOILERPLATE_PATTERNS
    if scheme == "strict_corpus":
        patterns += STRICT_BOILERPLATE_PATTERNS
    return patterns


def clean_text_with_log(text: str, scheme: str = "news_article") -> tuple[str, dict[str, Any]]:
    text = normalize_unicode_and_space(text)
    paragraphs = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
    cleaned: list[str] = []
    seen_short: set[str] = set()
    removed = {"boilerplate": 0, "duplicate_short": 0, "too_short": 0, "url_line": 0}
    compiled = [re.compile(p, re.I) for p in patterns_for_scheme(scheme)]
    strict = (scheme or "").lower() == "strict_corpus"
    light = (scheme or "").lower() == "light_archive"
    for p in paragraphs:
        p = re.sub(r"\s+", " ", p).strip()
        if not p:
            continue
        if not light and any(rx.search(p) for rx in compiled):
            removed["boilerplate"] += 1
            continue
        if strict and len(p) < 12:
            removed["too_short"] += 1
            continue
        if strict and re.fullmatch(r"https?://\S+|www\.\S+", p, flags=re.I):
            removed["url_line"] += 1
            continue
        key = p.lower()
        short_threshold = 160 if strict else 120
        if len(p) <= short_threshold:
            if key in seen_short:
                removed["duplicate_short"] += 1
                continue
            seen_short.add(key)
        cleaned.append(p)
    final = "\n\n".join(cleaned).strip()
    log = {
        "scheme": scheme,
        "input_paragraphs": len(paragraphs),
        "output_paragraphs": len(cleaned),
        "removed": removed,
    }
    return final, log


def clean_text(text: str, scheme: str = "news_article") -> str:
    return clean_text_with_log(text, scheme)[0]


def word_count(text: str) -> int:
    # Approximate corpus-size count across scripts.  Whitespace languages are
    # counted by word tokens; CJK scripts are counted by characters because they
    # are commonly unsegmented in raw web pages.
    text = text or ""
    cjk = len(re.findall(r"[\u4e00-\u9fff\u3040-\u30ff\u3400-\u4dbf\uac00-\ud7af]", text))
    tokens = len(re.findall(r"[A-Za-zÀ-ÖØ-öø-ÿĀ-žЀ-ӿΑ-ωΆ-ώ0-9][A-Za-zÀ-ÖØ-öø-ÿĀ-žЀ-ӿΑ-ωΆ-ώ0-9'’-]*", text))
    other = len(re.findall(r"[\u0600-\u06ff\u0900-\u097f\u0980-\u09ff\u0e00-\u0e7f]+", text))
    return cjk + tokens + other


def quality_score(text: str, meta: dict[str, str], status_code: int) -> float:
    score = 0
    chars = len(text or "")
    paras = len([p for p in re.split(r"\n\s*\n", text or "") if p.strip()])
    if 200 <= int(status_code or 0) < 300:
        score += 15
    if chars >= 800:
        score += 25
    elif chars >= 300:
        score += 15
    elif chars >= 100:
        score += 5
    if paras >= 4:
        score += 15
    elif paras >= 2:
        score += 8
    if meta.get("title"):
        score += 12
    if meta.get("published_time"):
        score += 10
    if meta.get("author"):
        score += 5
    low = (text or "").lower()
    noise_hits = sum(1 for key in ["cookie", "subscribe", "sign in", "advertisement", "privacy policy"] if key in low)
    score += max(0, 13 - noise_hits * 3)
    if meta.get("language"):
        score += 5
    return float(max(0, min(100, score)))


def append_jsonl(path: Path, data: dict[str, Any], lock: Optional[threading.Lock] = None) -> None:
    line = json.dumps(data, ensure_ascii=False) + "\n"
    if lock:
        with lock:
            with path.open("a", encoding="utf-8") as f:
                f.write(line)
    else:
        with path.open("a", encoding="utf-8") as f:
            f.write(line)


def metadata_excel_row(metadata: dict[str, Any], error: str = "") -> dict[str, Any]:
    meta = metadata.get("metadata", {}) if isinstance(metadata.get("metadata"), dict) else {}
    paths = metadata.get("paths", {}) if isinstance(metadata.get("paths"), dict) else {}
    counts = metadata.get("counts", {}) if isinstance(metadata.get("counts"), dict) else {}
    search = metadata.get("search_record", {}) if isinstance(metadata.get("search_record"), dict) else {}
    return {
        "processed_at": metadata.get("processed_at", ""),
        "ok": metadata.get("ok", True),
        "url": metadata.get("url", meta.get("url", "")),
        "final_url": metadata.get("final_url", ""),
        "canonical_url": metadata.get("canonical_url", meta.get("canonical_url", "")),
        "domain": metadata.get("domain", meta.get("domain", "")),
        "fetch_method": metadata.get("fetch_method", ""),
        "selenium_fallback_used": metadata.get("selenium_fallback_used", ""),
        "status_code": metadata.get("status_code", ""),
        "content_type": metadata.get("content_type", ""),
        "encoding": metadata.get("encoding", ""),
        "title": meta.get("title", ""),
        "author": meta.get("author", ""),
        "byline": meta.get("byline", ""),
        "publisher": meta.get("publisher", ""),
        "publication": meta.get("publication", ""),
        "organization": meta.get("organization", ""),
        "site_name": meta.get("site_name", ""),
        "source_name": meta.get("source_name", ""),
        "section": meta.get("section", ""),
        "keywords": meta.get("keywords", ""),
        "dateline": meta.get("dateline", ""),
        "location": meta.get("location", ""),
        "published_time": meta.get("published_time", ""),
        "modified_time": meta.get("modified_time", ""),
        "language": meta.get("language", ""),
        "description": meta.get("description", ""),
        "article_id": meta.get("article_id", ""),
        "copyright_holder": meta.get("copyright_holder", ""),
        "copyright_year": meta.get("copyright_year", ""),
        "is_accessible_for_free": meta.get("is_accessible_for_free", ""),
        "image_url": meta.get("image_url", ""),
        "publisher_logo": meta.get("publisher_logo", ""),
        "search_engine": search.get("search_engine", ""),
        "search_query": search.get("query", ""),
        "query_raw": search.get("query_raw", ""),
        "search_vertical": search.get("search_vertical", ""),
        "source_filter": search.get("source_filter", ""),
        "sort_mode": search.get("sort_mode", ""),
        "site_limit": search.get("site_limit", ""),
        "actual_domain": search.get("actual_domain", ""),
        "date_filter_type": search.get("date_filter_type", ""),
        "date_start": search.get("date_start", ""),
        "date_end": search.get("date_end", ""),
        "start_ts": search.get("start_ts", ""),
        "end_ts": search.get("end_ts", ""),
        "baidu_gpc": search.get("baidu_gpc", ""),
        "search_url": search.get("search_url", ""),
        "source": search.get("source", ""),
        "rank": search.get("rank", ""),
        "extraction_method": metadata.get("extraction_method", ""),
        "cleaning_scheme": metadata.get("cleaning_scheme", ""),
        "char_count": counts.get("chars", metadata.get("char_count", "")),
        "word_count": counts.get("words", metadata.get("word_count", "")),
        "paragraph_count": counts.get("paragraphs", ""),
        "quality_score": metadata.get("quality_score", ""),
        "raw_html_path": paths.get("raw_html", metadata.get("raw_html_path", "")),
        "raw_bytes_path": paths.get("raw_bytes", metadata.get("raw_bytes_path", "")),
        "raw_text_path": paths.get("raw_text", metadata.get("raw_text_path", "")),
        "clean_text_path": paths.get("clean_text", metadata.get("clean_text_path", "")),
        "metadata_path": paths.get("metadata", metadata.get("metadata_path", "")),
        "request_error": metadata.get("request_error", ""),
        "selenium_error": metadata.get("selenium_error", ""),
        "error": error or metadata.get("error", ""),
    }

def append_metadata_excel(path: Path, metadata: dict[str, Any], lock: Optional[threading.Lock] = None, error: str = "") -> None:
    def write() -> None:
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except Exception:
            return
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            if path.exists():
                wb = load_workbook(path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Content Metadata"
                ws.append(METADATA_EXCEL_FIELDS)
                fill = PatternFill("solid", fgColor="17384A")
                font = Font(color="FFFFFF", bold=True)
                for cell in ws[1]:
                    cell.fill = fill
                    cell.font = font
                    cell.alignment = Alignment(horizontal="center")
                ws.freeze_panes = "A2"
                from openpyxl.utils import get_column_letter
                for col_idx, field in enumerate(METADATA_EXCEL_FIELDS, start=1):
                    col = get_column_letter(col_idx)
                    if field in {"url", "final_url", "canonical_url", "search_url", "description"}:
                        width = 56
                    elif field.endswith("path") or field in {"baidu_gpc", "query_raw"}:
                        width = 46
                    elif field in {"title", "author", "publisher", "publication", "source_name"}:
                        width = 28
                    else:
                        width = 18
                    ws.column_dimensions[col].width = width
            row = metadata_excel_row(metadata, error=error)
            ws.append([row.get(f, "") for f in METADATA_EXCEL_FIELDS])
            for col_idx in range(1, len(METADATA_EXCEL_FIELDS) + 1):
                cell = ws.cell(ws.max_row, col_idx)
                cell.alignment = Alignment(vertical="top", wrap_text=False)
            wb.save(path)
        except Exception:
            # Metadata JSONL and per-page JSON remain the authoritative fallback.
            return

    if lock:
        with lock:
            write()
    else:
        write()


def write_manifest_and_excel(root: Path, metadata: dict[str, Any], lock: threading.Lock, error: str = "") -> tuple[Path, Path]:
    manifest_path = root / "content_manifest.jsonl"
    excel_path = root / "content_metadata.xlsx"
    with lock:
        with manifest_path.open("a", encoding="utf-8") as f:
            f.write(json.dumps(metadata, ensure_ascii=False) + "\n")
        append_metadata_excel(excel_path, metadata, None, error=error)
    return manifest_path, excel_path


def save_html_content_result(
    record: Any,
    html_text: str,
    final_url: str,
    status_code: int,
    content_type: str,
    encoding: str,
    settings: ContentDownloadSettings,
    index: int,
    manifest_lock: threading.Lock,
    fetch_method: str = "requests",
    request_error: str = "",
    selenium_error: str = "",
    selenium_fallback_used: bool = False,
    raw_bytes: bytes | None = None,
) -> ContentResult:
    dirs = ensure_content_dirs(settings.content_root)
    url = record_url(record)
    domain = domain_of(final_url or url)
    stem = cache_stem(record, index)
    raw_html_path = dirs["raw_html"] / f"{stem}.html"
    raw_bytes_path = dirs["raw_html"] / f"{stem}.raw"
    raw_text_path = dirs["raw_text"] / f"{stem}.txt"
    clean_text_path = dirs["clean_text"] / f"{stem}.txt"
    metadata_path = dirs["metadata"] / f"{stem}.json"

    html_text = repair_mojibake_text(html_text or "")
    raw_html_path.write_text(html_text, encoding="utf-8", errors="ignore")
    raw_bytes_path_str = ""
    if raw_bytes is not None:
        try:
            raw_bytes_path.write_bytes(raw_bytes)
            raw_bytes_path_str = str(raw_bytes_path)
        except Exception:
            raw_bytes_path_str = ""

    # Metadata and text extraction must be fault-tolerant.  Some Chinese news sites
    # expose unusual JSON-LD/meta structures or JS-rendered DOM fragments; Selenium may
    # load the page correctly while a parser still raises errors such as
    # "NoneType object has no attribute 'get'".  In that case we still save the raw
    # HTML and fall back to a conservative BeautifulSoup visible-text extraction.
    extraction_errors: list[str] = []
    try:
        meta = extract_metadata(html_text, final_url or url)
        if not isinstance(meta, dict):
            meta = {}
    except Exception as exc:
        meta = {
            "domain": domain,
            "url": final_url or url,
            "metadata_error": str(exc),
            "metadata_traceback": traceback.format_exc(limit=5),
        }
        extraction_errors.append(f"metadata: {exc}")

    try:
        raw_text, extraction_method, resolved_scheme, extra_meta = extract_text_by_scheme(
            html_text, final_url or url, record, settings.cleaning_scheme
        )
        if not isinstance(extra_meta, dict):
            extra_meta = {}
    except Exception as exc:
        extraction_errors.append(f"text_extraction: {exc}")
        try:
            raw_text = extract_bs4_article_text(html_text, remove_layout=True)
            extraction_method = "bs4_news_article_emergency_fallback"
            resolved_scheme = "news_article"
            extra_meta = {"text_extraction_error": str(exc), "text_extraction_traceback": traceback.format_exc(limit=5)}
        except Exception as exc2:
            extraction_errors.append(f"emergency_bs4: {exc2}")
            soup_fallback = BeautifulSoup(html_text or "", "html.parser")
            raw_text = soup_fallback.get_text("\n", strip=True)
            extraction_method = "visible_text_last_resort"
            resolved_scheme = "light_archive"
            extra_meta = {
                "text_extraction_error": str(exc),
                "emergency_extraction_error": str(exc2),
                "text_extraction_traceback": traceback.format_exc(limit=5),
            }

    for k, v in (extra_meta or {}).items():
        if v and not meta.get(k):
            meta[k] = v
    if extraction_errors:
        meta["extraction_warnings"] = " | ".join(extraction_errors)

    try:
        cleaned, clean_log = clean_text_with_log(raw_text, resolved_scheme)
    except Exception as exc:
        cleaned = normalize_unicode_and_space(raw_text)
        clean_log = {"scheme": resolved_scheme, "error": str(exc), "fallback": "raw_normalized"}
    raw_text_path.write_text(raw_text or "", encoding="utf-8")
    clean_text_path.write_text(cleaned or "", encoding="utf-8")

    wc = word_count(cleaned)
    para_count = len([p for p in cleaned.split("\n\n") if p.strip()])
    score = quality_score(cleaned, meta, status_code)
    metadata = {
        "processed_at": utc_now_iso(),
        "ok": True,
        "url": url,
        "final_url": final_url or url,
        "canonical_url": meta.get("canonical_url", ""),
        "domain": domain,
        "fetch_method": fetch_method,
        "selenium_fallback_used": bool(selenium_fallback_used),
        "request_error": request_error,
        "selenium_error": selenium_error,
        "status_code": status_code,
        "content_type": content_type,
        "encoding": encoding or "",
        "search_record": record_dict(record),
        "metadata": meta,
        "paths": {
            "raw_html": str(raw_html_path),
            "raw_bytes": raw_bytes_path_str,
            "raw_text": str(raw_text_path),
            "clean_text": str(clean_text_path),
            "metadata": str(metadata_path),
        },
        "counts": {"chars": len(cleaned), "words": wc, "paragraphs": para_count},
        "quality_score": score,
        "extraction_method": extraction_method,
        "cleaning_scheme": resolved_scheme,
        "clean_log": clean_log,
    }
    metadata_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    manifest_path, excel_path = write_manifest_and_excel(dirs["root"], metadata, manifest_lock)

    return ContentResult(
        ok=True,
        url=url,
        final_url=final_url or url,
        domain=domain,
        status_code=status_code,
        content_type=content_type,
        encoding=encoding or "",
        title=meta.get("title", ""),
        author=meta.get("author", ""),
        published_time=meta.get("published_time", ""),
        language=meta.get("language", ""),
        extraction_method=extraction_method,
        cleaning_scheme=resolved_scheme,
        char_count=len(cleaned),
        word_count=wc,
        quality_score=score,
        raw_html_path=str(raw_html_path),
        raw_text_path=str(raw_text_path),
        clean_text_path=str(clean_text_path),
        metadata_path=str(metadata_path),
        metadata_excel_path=str(excel_path),
        manifest_path=str(manifest_path),
        fetch_method=fetch_method,
        selenium_fallback_used=bool(selenium_fallback_used),
    )


def save_content_result(record: Any, response: requests.Response, settings: ContentDownloadSettings, index: int, manifest_lock: threading.Lock) -> ContentResult:
    html_text, chosen_encoding, _decode_note = decode_response_html(response)
    return save_html_content_result(
        record=record,
        html_text=html_text,
        final_url=response.url or record_url(record),
        status_code=response.status_code,
        content_type=response.headers.get("Content-Type", ""),
        encoding=chosen_encoding or getattr(response, "encoding", "") or "",
        settings=settings,
        index=index,
        manifest_lock=manifest_lock,
        fetch_method="requests",
        raw_bytes=getattr(response, "content", b"") or b"",
    )

def save_non_html_result(record: Any, response: requests.Response, settings: ContentDownloadSettings, index: int, manifest_lock: threading.Lock, error: str) -> ContentResult:
    dirs = ensure_content_dirs(settings.content_root)
    stem = cache_stem(record, index)
    ctype = response.headers.get("Content-Type", "").lower()
    suffix = ".bin"
    if "pdf" in ctype:
        suffix = ".pdf"
    elif "text" in ctype:
        suffix = ".txt"
    raw_path = dirs["raw_html"] / f"{stem}{suffix}"
    raw_path.write_bytes(response.content)
    metadata = {
        "processed_at": utc_now_iso(),
        "ok": False,
        "url": record_url(record),
        "final_url": response.url or record_url(record),
        "domain": domain_of(response.url or record_url(record)),
        "fetch_method": "requests",
        "selenium_fallback_used": False,
        "request_error": "",
        "selenium_error": "",
        "status_code": response.status_code,
        "content_type": response.headers.get("Content-Type", ""),
        "encoding": response.encoding or "",
        "search_record": record_dict(record),
        "metadata": {"domain": domain_of(response.url or record_url(record))},
        "paths": {"raw_html": str(raw_path)},
        "counts": {"chars": 0, "words": 0, "paragraphs": 0},
        "quality_score": 0,
        "extraction_method": "unsupported_non_html",
        "cleaning_scheme": settings.cleaning_scheme,
        "error": error,
    }
    manifest_path, excel_path = write_manifest_and_excel(dirs["root"], metadata, manifest_lock, error=error)
    return ContentResult(
        ok=False,
        url=record_url(record),
        final_url=response.url or record_url(record),
        domain=metadata["domain"],
        status_code=response.status_code,
        content_type=response.headers.get("Content-Type", ""),
        raw_html_path=str(raw_path),
        metadata_excel_path=str(excel_path),
        manifest_path=str(manifest_path),
        error=error,
    )


def save_error_result(record: Any, settings: ContentDownloadSettings, manifest_lock: threading.Lock, error: str) -> ContentResult:
    dirs = ensure_content_dirs(settings.content_root)
    url = record_url(record)
    dom = domain_of(url)
    metadata = {
        "processed_at": utc_now_iso(),
        "ok": False,
        "url": url,
        "final_url": "",
        "domain": dom,
        "fetch_method": "requests",
        "selenium_fallback_used": False,
        "request_error": error,
        "selenium_error": "",
        "status_code": 0,
        "content_type": "",
        "encoding": "",
        "search_record": record_dict(record),
        "metadata": {"domain": dom},
        "paths": {},
        "counts": {"chars": 0, "words": 0, "paragraphs": 0},
        "quality_score": 0,
        "extraction_method": "download_error",
        "cleaning_scheme": settings.cleaning_scheme,
        "error": error,
    }
    manifest_path, excel_path = write_manifest_and_excel(dirs["root"], metadata, manifest_lock, error=error)
    return ContentResult(ok=False, url=url, domain=dom, metadata_excel_path=str(excel_path), manifest_path=str(manifest_path), error=error)


SELENIUM_BLOCKED_PATTERNS = re.compile(
    r"captcha|verify you are human|human verification|access denied|just a moment|checking your browser|enable javascript|unusual traffic|robot check|cf-browser-verification",
    re.I,
)

_selenium_fallback_lock = threading.Lock()


def html_needs_selenium_fallback(html_text: str, status_code: int, content_type: str, record: Any, settings: ContentDownloadSettings) -> bool:
    if not bool(getattr(settings, "selenium_fallback", True)):
        return False
    if int(status_code or 0) in {401, 403, 407, 408, 409, 425, 429, 500, 502, 503, 504}:
        return True
    text = html_text or ""
    low = text[:20000].lower()
    if SELENIUM_BLOCKED_PATTERNS.search(low):
        return True
    ctype = (content_type or "").lower()
    if "html" in ctype or "xml" in ctype or text.lstrip().startswith("<"):
        # If the static HTML produces almost no article text, the page may be JS-rendered.
        try:
            raw_text, _method, _scheme, _meta = extract_text_by_scheme(text, record_url(record), record, settings.cleaning_scheme)
            if not is_usable_text(raw_text, min_chars=180) and len(text) < 120000:
                visible = BeautifulSoup(text, "html.parser").get_text(" ", strip=True)
                if len(visible) < 800 or re.search(r"__NEXT_DATA__|window\.__|id=[\"']root[\"']|id=[\"']app[\"']", text, re.I):
                    return True
        except Exception:
            return True
    return False


def fetch_with_selenium_fallback(url: str, settings: ContentDownloadSettings) -> tuple[str, str, str]:
    """Fetch rendered HTML with Selenium as an optional fallback for content pages.

    A fresh browser is opened for the single URL and closed immediately.  The module-level
    lock intentionally serializes Selenium fallback sessions so multiple worker threads do
    not start several heavy browsers at once.
    """
    with _selenium_fallback_lock:
        backend = (getattr(settings, "selenium_backend", "selenium_chrome") or "selenium_chrome").lower()
        driver = None
        try:
            from selenium import webdriver  # type: ignore
            from selenium.webdriver.chrome.options import Options as ChromeOptions  # type: ignore
            from selenium.webdriver.chrome.service import Service as ChromeService  # type: ignore
            from selenium.webdriver.edge.options import Options as EdgeOptions  # type: ignore
            from selenium.webdriver.edge.service import Service as EdgeService  # type: ignore
        except Exception as exc:
            raise RuntimeError("Selenium fallback requested, but selenium is not installed.") from exc
        try:
            if backend == "selenium_edge":
                options = EdgeOptions()
                if getattr(settings, "selenium_headless", False):
                    options.add_argument("--headless=new")
                if getattr(settings, "selenium_binary_path", ""):
                    options.binary_location = str(settings.selenium_binary_path)
                options.add_argument(f"--user-agent={settings.user_agent or USER_AGENT_FALLBACK}")
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("--disable-notifications")
                options.add_argument("--lang=en-US")
                driver_path = str(getattr(settings, "selenium_driver_path", "") or "").strip()
                if driver_path and Path(driver_path).exists():
                    driver = webdriver.Edge(service=EdgeService(executable_path=driver_path), options=options)
                else:
                    driver = webdriver.Edge(options=options)
            else:
                options = ChromeOptions()
                if getattr(settings, "selenium_headless", False):
                    options.add_argument("--headless=new")
                if getattr(settings, "selenium_binary_path", ""):
                    options.binary_location = str(settings.selenium_binary_path)
                options.add_argument(f"--user-agent={settings.user_agent or USER_AGENT_FALLBACK}")
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("--disable-notifications")
                options.add_argument("--lang=en-US")
                driver_path = str(getattr(settings, "selenium_driver_path", "") or "").strip()
                if driver_path and Path(driver_path).exists():
                    driver = webdriver.Chrome(service=ChromeService(executable_path=driver_path), options=options)
                else:
                    driver = webdriver.Chrome(options=options)
            hard_timeout = max(5, int(getattr(settings, "timeout_seconds", 30) or 30))
            driver.set_page_load_timeout(hard_timeout)
            try:
                driver.set_script_timeout(hard_timeout)
            except Exception:
                pass
            driver.get(url)
            wait_seconds = max(0.5, int(getattr(settings, "selenium_wait_ms", 3500) or 3500) / 1000.0)
            time.sleep(wait_seconds)
            try:
                driver.execute_script("return document.readyState")
            except Exception:
                pass
            final_url = getattr(driver, "current_url", "") or url
            page_html = repair_mojibake_text(getattr(driver, "page_source", "") or "")
            if not page_html.strip():
                raise RuntimeError("Selenium returned an empty page source.")
            return page_html, final_url, backend
        finally:
            try:
                if driver is not None:
                    driver.quit()
            except Exception:
                pass


def download_one(record: Any, settings: ContentDownloadSettings, index: int, domain_locks: DomainLockPool, manifest_lock: threading.Lock, stop_checker=None) -> ContentResult:
    url = record_url(record)
    if not url:
        return ContentResult(ok=False, url="", error="Empty URL")
    normalized = normalize_url(url)
    dom = domain_of(normalized)
    headers = {
        "User-Agent": settings.user_agent or USER_AGENT_FALLBACK,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.8,zh-CN;q=0.7,zh;q=0.6,*;q=0.5",
        "Connection": "close",
    }
    request_error = ""
    selenium_error = ""
    fetch_mode = (getattr(settings, "fetch_mode", "mixed") or "mixed").lower().strip()
    if fetch_mode not in {"requests", "selenium", "mixed"}:
        fetch_mode = "mixed"
    receive_wait_ms = max(0, int(getattr(settings, "receive_wait_ms", 0) or 0))
    try:
        lock = domain_locks.lock_for(dom)
        lock_timeout = max(1, int(getattr(settings, "domain_lock_timeout_seconds", 300) or 300))
        acquired = lock.acquire(timeout=lock_timeout)
        if not acquired:
            return save_error_result(record, settings, manifest_lock, f"Domain lock timeout after {lock_timeout} seconds for {dom}")
        try:
            mn = max(0, int(settings.min_delay_ms or 0))
            mx = max(mn, int(settings.max_delay_ms or mn))
            if mx > 0:
                slept = 0.0
                target = random.uniform(mn, mx) / 1000.0
                while slept < target:
                    if stop_checker and stop_checker():
                        return ContentResult(ok=False, url=url, domain=dom, error="Stopped by user")
                    step = min(0.5, target - slept)
                    time.sleep(step)
                    slept += step
            if stop_checker and stop_checker():
                return ContentResult(ok=False, url=url, domain=dom, error="Stopped by user")

            if fetch_mode == "selenium":
                try:
                    html_text, final_url, backend = fetch_with_selenium_fallback(normalized, settings)
                    return save_html_content_result(
                        record, html_text, final_url, 200, "text/html; selenium-rendered", "utf-8",
                        settings, index, manifest_lock, fetch_method=backend,
                        request_error="", selenium_fallback_used=False,
                    )
                except Exception as exc:
                    return save_error_result(record, settings, manifest_lock, str(exc))

            resp: Optional[requests.Response] = None
            try:
                resp = requests.get(normalized, headers=headers, timeout=max(1, int(settings.timeout_seconds or 30)), allow_redirects=True)
                if receive_wait_ms > 0:
                    time.sleep(receive_wait_ms / 1000.0)
            except Exception as exc:
                request_error = str(exc)

            # If requests failed entirely, try Selenium before recording failure.
            if resp is None:
                if fetch_mode == "mixed" and bool(getattr(settings, "selenium_fallback", True)):
                    try:
                        html_text, final_url, backend = fetch_with_selenium_fallback(normalized, settings)
                        return save_html_content_result(
                            record, html_text, final_url, 200, "text/html; selenium-rendered", "utf-8",
                            settings, index, manifest_lock, fetch_method=f"{backend}_fallback",
                            request_error=request_error, selenium_fallback_used=True,
                        )
                    except Exception as exc:
                        selenium_error = str(exc)
                return save_error_result(record, settings, manifest_lock, request_error + (f" | Selenium fallback failed: {selenium_error}" if selenium_error else ""))

            decoded_text, decoded_encoding, _decode_note = decode_response_html(resp)
            ctype = resp.headers.get("Content-Type", "").lower()
            htmlish = ("html" in ctype or "xml" in ctype or (decoded_text or "").lstrip().startswith("<"))

            # Try Selenium if the HTTP status or the static HTML suggests blocking or JS-only rendering.
            if fetch_mode == "mixed" and bool(getattr(settings, "selenium_fallback", True)) and htmlish and html_needs_selenium_fallback(decoded_text or "", resp.status_code, ctype, record, settings):
                try:
                    html_text, final_url, backend = fetch_with_selenium_fallback(resp.url or normalized, settings)
                    # Prefer Selenium only when it yields a usable article or when requests clearly failed.
                    try:
                        probe_text, _m, _sch, _em = extract_text_by_scheme(html_text, final_url, record, settings.cleaning_scheme)
                        selenium_usable = is_usable_text(probe_text, min_chars=180)
                    except Exception:
                        selenium_usable = len(BeautifulSoup(html_text, "html.parser").get_text(" ", strip=True)) >= 300
                    if selenium_usable or int(resp.status_code or 0) >= 400:
                        return save_html_content_result(
                            record, html_text, final_url, 200 if int(resp.status_code or 0) >= 400 else resp.status_code,
                            "text/html; selenium-rendered", "utf-8", settings, index, manifest_lock,
                            fetch_method=f"{backend}_fallback", request_error="" if resp.ok else f"HTTP {resp.status_code}",
                            selenium_fallback_used=True,
                        )
                except Exception as exc:
                    selenium_error = str(exc)
                    # Keep the requests result if it is at least available.

            if int(resp.status_code or 0) >= 400:
                if selenium_error:
                    return save_error_result(record, settings, manifest_lock, f"HTTP {resp.status_code}; Selenium fallback failed: {selenium_error}")
                return save_error_result(record, settings, manifest_lock, f"HTTP {resp.status_code}")

            if not htmlish:
                return save_non_html_result(
                    record, resp, settings, index, manifest_lock,
                    "Downloaded, but content type is not HTML; text extraction is not implemented for this type yet.",
                )
            result = save_html_content_result(
                record=record, html_text=decoded_text, final_url=resp.url or normalized,
                status_code=resp.status_code, content_type=resp.headers.get("Content-Type", ""),
                encoding=decoded_encoding, settings=settings, index=index, manifest_lock=manifest_lock,
                fetch_method="requests", raw_bytes=getattr(resp, "content", b"") or b"",
            )
            if selenium_error and result.ok:
                # The page was still processed through requests; keep fallback failure evidence in the JSON/Excel on future runs through manifest.
                pass
            return result
        finally:
            try:
                lock.release()
            except Exception:
                pass
    except Exception as exc:
        try:
            return save_error_result(record, settings, manifest_lock, str(exc))
        except Exception:
            return ContentResult(ok=False, url=url, domain=dom, error=str(exc))
